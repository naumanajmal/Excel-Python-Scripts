import pyautogui
import time
import openpyxl as xl
import clipboard
wb = xl.load_workbook('Book1.xlsx')
sheet = wb['Sheet1']
hadithnumber = 21
nexthadithno = hadithnumber + 1
hadithnumber = str(hadithnumber)
nexthadithno = str(nexthadithno)
data = ""
track = 0
for row in range(28, 890):
  cell1 = sheet.cell(row, 1)
  cell2 = sheet.cell(row, 2)
  cell3 = sheet.cell(row, 3)
  if track > 5:
      break
  if data == "":
   pyautogui.click(227, 314)
   pyautogui.hotkey('ctrl', 'a')
   time.sleep(1.5)
   pyautogui.hotkey('ctrl', 'c')
   pyautogui.click(316, 74)
   time.sleep(0.5)
   data = clipboard.paste()
   time.sleep(0.5)
  n = len(hadithnumber)
  if hadithnumber in data and not data[:n] == hadithnumber:

     hadithtext = data.partition(hadithnumber)[0]
     data = data.partition(hadithnumber)[2]
     data = hadithnumber + data
     cell2.value = hadithtext
     track = 0
     print(track)
     print("previous  text")
     continue
  if hadithnumber not in data and nexthadithno not in data:
    cell2.value = data
    data = ""
    track =  track +1
    print(track)
    print("only previoius text")
    continue
  if hadithnumber in data and data[:n] == hadithnumber and nexthadithno in data:
    hadithtext = data.partition(nexthadithno)[0]
    data = data.partition(nexthadithno)[2]
    data = nexthadithno + data
    cell1.value = hadithtext
    hadithnumber = nexthadithno
    nexthadithno = int(hadithnumber) + 1
    nexthadithno = str (nexthadithno)
    track = 0
    print(track)
    print("single hadith with next text")
    continue
  if hadithnumber in data and data[:n] == hadithnumber and nexthadithno not in data:
    hadithtext = data
    hadithnumber = str(int(hadithnumber) + 1)
    nexthadithno = str(int(hadithnumber) + 1)
    data = ""
    track = 0
    print(track)
    cell1.value = hadithtext
    print("single hadith")
    continue
  """if hadithnumber not in data and nexthadithno in data and data[:n] == hadithnumber:
    hadithtext = data
    hadithnumber = str(int(nexthadithno) + 1)
    nexthadithno = str(int(hadithnumber) + 1)
    data = ""
    track = 0
    cell1.value = hadithtext
    continue
  if hadithnumber not in data and nexthadithno in data:
      hadithtext = data.partition(nexthadithno)[0]
      data = data.partition(nexthadithno)[2]
      data = nexthadithno + data
      cell1.value = hadithtext
      track = 0
      hadithnumber = str(int(nexthadithno) +1)
      nexthadithno = str(int(hadithnumber) + 1)
      continue"""
wb.save('Book1.xlsx')

