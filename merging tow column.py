import pyautogui
import time
import openpyxl as xl
import clipboard
wb = xl.load_workbook('Book1.xlsx')
sheet = wb['Sheet1']

for row in range(2, 2668):
  cell2 = sheet.cell(row, 2)
  cell3 = sheet.cell(row, 3)
  cell4 = sheet.cell(row, 4)
  data = str(cell2.value)
  dat1 = str(cell3.value)
  cell4.value = data + dat1

wb.save('Book1.xlsx')

