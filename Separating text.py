import pyautogui
import time
import openpyxl as xl
import clipboard
wb = xl.load_workbook('Book1.xlsx')
sheet = wb['Sheet1']

for row in range(2, 874):
  cell1 = sheet.cell(row, 13)
  cell3 = sheet.cell(row, 14)
  data = str(cell1.value)
  word = "موضوع"
  if word in data:
    cell3.value = "mozu"


wb.save('Book1.xlsx')

