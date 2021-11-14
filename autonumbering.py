import pyautogui
import time
import openpyxl as xl
import clipboard
wb = xl.load_workbook('Book1.xlsx')
sheet = wb['Sheet1']
previoustext = ""
number = 0
for row in range(2, 231):
  cell1 = sheet.cell(row, 5)
  cell2 = sheet.cell(row, 4)
  text = str(cell1.value)
  if previoustext != text:
   previoustext = text
   number = number +1
   cell2.value = number
  else:
   cell2.value = number


wb.save('Book1.xlsx')

