import pyautogui
import time
import openpyxl as xl
import clipboard
wb = xl.load_workbook('Book1.xlsx')
sheet = wb['Sheet1']
celltext = ""
for row in range(2, 231):
  cell1 = sheet.cell(row, 4)
  cell2 = sheet.cell(row, 5)
  text = str(cell1.value)
  print(cell1.value)
  if cell1.value != None:
   celltext = text
   cell2.value = celltext
  else:
   cell2.value = celltext


wb.save('Book1.xlsx')

