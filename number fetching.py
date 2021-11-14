import pyautogui
import time
import openpyxl as xl
import clipboard
wb = xl.load_workbook('Book1.xlsx')
sheet = wb['Sheet1']
final = ""
for row in range(2, 2919):
  cell1 = sheet.cell(row, 1)
  cell3 = sheet.cell(row, 2)
  data = str(cell1.value)
  datacut = data[:5]
  for element in datacut:
       if element.isdigit():
           final = final + element
           if element == "/":
               break;

  cell3.value = final
  final = ""



wb.save('Book1.xlsx')

